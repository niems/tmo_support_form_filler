import openpyxl
from Log import Log
import sys

class ExcelCommands(object):
    #determines range to search through in sheet
    def sheet_range(ws, col_range, row_range): #internal only
        try:
            if type(col_range) is not type(None):
                if col_range[0] is None: #use full column range
                    col_range[0] = 1

                if col_range[1] is None:
                    col_range[1] = ws.columns + 1
            else:
                col_range = [1, ws.columns + 1]

            if type(row_range) is not type(None):
                if row_range[0] is None: #use full row range
                    row_range[0] = 1

                if row_range[1] is None:
                    row_range[1] = len( list(ws.rows) ) + 1
            else:
                row_range = [1, ws.rows]

        except Exception as e:
            print(e)
            print('err on line {}'.format(sys.exc_info[-1].tb_lineno))
            Log.pause()

        return col_range, row_range

    #internal only
    def does_value_exist(ws, col_range, row_range, key):
        pos = None
        for row in range(row_range[0], row_range[1] + 1):
            for col in range(col_range[0], col_range[1] + 1):
                if ws.cell(row = row, column = col).value == key:
                    pos = [row, col] #key position
                    break

        return pos #key not found

    def get_cell_value(ws, col = 0, row = 0):
        return ws.cell(row = row, column = col).value

    def find_value(ws, col_range = [], row_range = [], key = None):
        try:
            current_pos = [0, 0] #position of value to find

            if col_range[0] == -1:
                col_range[0] = len( list(ws.columns) ) + 1

            if col_range[1] == -1:
                col_range[1] = len( list(ws.columns) ) + 1

            if row_range[0] == -1:
                row_range[0] = len( list(ws.rows) ) + 1

            if row_range[1] == -1:
                row_range[1] = len( list(ws.rows) ) + 1

            col_range, row_range = ExcelCommands.sheet_range(ws, col_range, row_range)
            #print('find val row range: {}'.format(row_range) )
            current_pos = ExcelCommands.does_value_exist(ws, col_range, row_range, key)

            if current_pos is None:
                pass#print('current position is None.')

        except Exception as e:
            print('{}'.format(e.with_traceback() ) )
            Log.pause()

        return current_pos #position key was found in sheet

    #returns None if not found, [row,col] of value if found
    def find_value(wb_name, ws = None, col_range = [],
                   row_range = [], key = None):
        try:
            wb = openpyxl.load_workbook(wb_name)
            current_pos = [0, 0] #position of value to find

            if ws is None: #defaults to active sheet
                ws = wb.active

            else: #code to pull specified sheet if it exists. If false, default to active sheet
                pass

            if col_range[0] == -1:
                col_range[0] = len( list(ws.columns) ) + 1

            if col_range[1] == -1:
                col_range[1] = len( list(ws.columns) ) + 1

            if row_range[0] == -1:
                row_range[0] = len( list(ws.rows) ) + 1

            if row_range[1] == -1:
                row_range[1] = len( list(ws.rows) ) + 1

            col_range, row_range = ExcelCommands.sheet_range(ws, col_range, row_range)
            #print('find val row range: {}'.format(row_range) )
            current_pos = ExcelCommands.does_value_exist(ws, col_range, row_range, key)

            if current_pos is None:
                pass#print('current position is None.')

        except Exception as e:
            print(e)
            print('err on line {}'.format(sys.exc_info[-1].tb_lineno))
            Log.pause()

        return current_pos #position key was found in sheet

    def is_value_in_sheet(ws, col_range = [-1, -1], row_range = [-1, -1], key = None):
        try:
            current_pos = [0, 0, 0] #position of value to find, and value of cell

            if col_range[0] == -1:
                col_range[0] = len( list(ws.columns) ) + 1

            if col_range[1] == -1:
                col_range[1] = len( list(ws.columns) ) + 1

            if row_range[0] == -1:
                row_range[0] = len( list(ws.rows) ) + 1

            if row_range[1] == -1:
                row_range[1] = len( list(ws.rows) ) + 1


            col_range, row_range = ExcelCommands.sheet_range(ws, col_range, row_range)
            #print('find row range: {}'.format(row_range) )
            current_pos = ExcelCommands.does_value_exist(ws, col_range, row_range, key)

        except Exception as e:
            print('{}'.format(e.with_traceback) )
            Log.pause()

        return current_pos #position key was found in sheet

    def add_value_to_sheet(wb_name, ws = None, pos = [], value = None): #pos = [row,col]
        try:
            wb = openpyxl.load_workbook(wb_name)

            if ws is None: #defaults to active sheet
                ws = wb.active

            else: #code to pull specified sheet if it exists. If false, default to active sheet
                pass

            if pos[0] == -1: #uses next new row
                pos[0] = len( list(ws.rows) ) + 1
                #print('current number of rows: {}'.format(pos[0]) )

            if pos[1] == -1: #uses next new column
                pos[1] = len( list(ws.columns) ) + 1

            ws.cell(row = pos[0], column = pos[1]).value = value
            #print('cell val: {}'.format( ws.cell(row = pos[0], column = pos[1]).value ) )
            wb.save(wb_name)

        except Exception as e:
            print(e)
            print('err on line {}'.format(sys.exc_info[-1].tb_lineno))
            Log.pause()

        return pos
